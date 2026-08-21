# -*- coding: utf-8 -*-
"""
build_tiktok_analysis.py  —  สร้างชีต "เงินเข้าบริษัท" และ "ใบปะหน้า TIKTOK"
Usage (CLI):   python build_tiktok_analysis.py <Data_ERP_YYMM.xlsx>
Usage (import): from scripts.build_tiktok_analysis import main; main("Data ERP 2605.xlsx")
"""
import sys
import os
import re as _re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


WANTED_COLS = [
    ('ยอดขาย',                                       'ยอดรวมค่าสินค้าหลังหักส่วนลดจากผู้ขาย'),
    ('ค่าธรรมเนียมคำสั่งซื้อ',                        'ค่าธรรมเนียมคำสั่งซื้อ'),
    ('ค่าคอมมิชชั่น TikTok Shop',                    'ค่าคอมมิชชั่น TikTok Shop'),
    ('ค่าธรรมเนียมการจัดส่งจริง',                    'ค่าธรรมเนียมการจัดส่งจริง'),
    ('ส่วนลดค่าธรรมเนียมการจัดส่งจากแพลตฟอร์ม',    'ส่วนลดค่าธรรมเนียมการจัดส่งจากแพลตฟอร์ม'),
    ('ค่าธรรมเนียมการจัดส่งของลูกค้า',              'ค่าธรรมเนียมการจัดส่งของลูกค้า'),
    ('เงินสนับสนุนการจัดส่ง',                        'เงินสนับสนุนการจัดส่ง'),
    ('ยอดรวมเงินคืนหลังหักส่วนลดจากผู้ขาย',         'ยอดรวมเงินคืนหลังหักส่วนลดจากผู้ขาย'),
    ('ค่าธรรมเนียมการจัดส่งสินค้าคืนตามจริง',       'ค่าธรรมเนียมการจัดส่งสินค้าคืนตามจริง'),
    ('เงินคืนสำหรับค่าจัดส่ง',                       'เงินคืนสำหรับค่าจัดส่ง'),
    ('ค่าธรรมเนียมสนับสนุนการเติบโตของร้านค้า',     'ค่าธรรมเนียมสนับสนุนการเติบโตของร้านค้า'),
    ('เงินเข้าบริษัท',                                'จำนวนเงินที่ชำระทั้งหมด'),
]

EXTRA_COLS = [
    'ต้นทุนสินค้า/ต่อชิ้น',
    'ต้นทุนสินค้ารวม',
    'กำไร(บาท)',
    'กำไร(%)',
    'หมายเหตุ',
    'สาเหตุติดลบ',
]

COST_URL  = 'https://docs.google.com/spreadsheets/d/1whwXYiUyDC0uo6oaudr_QqzCV6UFOWd0mL4Y2El9JRY/edit?usp=sharing'
STOCK_URL = 'https://docs.google.com/spreadsheets/d/1p4-DX-Sq-Mvo_FNAkYtGMkdEBvPDCuPM_sZ_b-_Tm-Y/edit?usp=sharing'


def col_letter_to_num(letter):
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n


def main(input_path: str, log_fn=None):
    """
    เพิ่มชีท 'เงินเข้าบริษัท' และ 'ใบปะหน้า TIKTOK' เข้าไฟล์เดิม
    input_path: path ของไฟล์ xlsx ที่สร้างจาก build_tiktok_erp
    log_fn: callable(str) สำหรับแสดง log ระหว่างประมวลผล
    """
    def log(msg):
        print(msg)
        if log_fn:
            log_fn(msg)

    OUTPUT = input_path

    # ── 1. อ่าน Data ERP ──────────────────────────────────────────────
    log('Reading Data ERP...')
    wb0 = load_workbook(input_path, data_only=True)
    ws0 = wb0['Data ERP']
    erp_headers = [c.value for c in ws0[1]]
    N = len(erp_headers)

    def find_idx(kws):
        for i, h in enumerate(erp_headers):
            for k in kws:
                if k in str(h or ''):
                    return i
        return None

    shop_i  = find_idx(['พนักงานขาย'])
    order_i = 5   # col F = TikTok order# (0-based)

    F_COL    = get_column_letter(order_i + 1)   # 'F'
    SHOP_COL = get_column_letter(shop_i + 1) if shop_i is not None else 'W'

    rows = [
        list(r) for r in ws0.iter_rows(min_row=2, values_only=True)
        if any(str(r[shop_i] or '').startswith(p) for p in ('TikTok-', 'Tiktok-'))
    ]
    log(f'  TikTok ERP rows: {len(rows)}')
    wb0.close()

    # ── 2. อ่าน TikTok sheet headers ─────────────────────────────────
    log('Reading TikTok headers...')
    wb1 = load_workbook(input_path, data_only=True)
    ws_tt_ref = wb1['TikTok']
    tt_hdrs = {
        str(c.value).strip(): get_column_letter(c.column)
        for c in ws_tt_ref[1]
        if c.value
    }
    TT_MR = ws_tt_ref.max_row
    wb1.close()

    tt_col_map = {}
    for display_name, tiktok_name in WANTED_COLS:
        if tiktok_name in tt_col_map:
            continue
        for h, col_letter in tt_hdrs.items():
            if tiktok_name == h or tiktok_name in h or h in tiktok_name:
                tt_col_map[tiktok_name] = col_letter
                break
        if tiktok_name not in tt_col_map:
            log(f'  WARNING: ไม่พบคอลัมน์: {tiktok_name}')

    TT_ORDER = tt_hdrs.get('หมายเลขคำสั่งซื้อ/การปรับ', 'B')
    TT_END   = max(TT_MR + 100, 3000)
    CE       = max(len(rows) + 200, 1000)

    log(f'  TT_ORDER={TT_ORDER}  TT_END={TT_END}')

    # ── 3. เปิด workbook ──────────────────────────────────────────────
    log('Opening workbook...')
    wb = load_workbook(input_path)

    # ── 3.5 แปลง TikTok financial columns: text → number ─────────────
    log('Converting TikTok numeric columns (text → number)...')
    ws_tt_fix = wb['TikTok']
    numeric_col_nums = {col_letter_to_num(cl) for cl in tt_col_map.values()}

    converted = 0
    for row in ws_tt_fix.iter_rows(min_row=2):
        for cell in row:
            if cell.column in numeric_col_nums and cell.value is not None:
                try:
                    num = float(str(cell.value).replace(',', '').strip())
                    cell.value = num
                    converted += 1
                except (ValueError, TypeError):
                    pass
    log(f'  Converted {converted} cells to number')

    # ── 4. สร้าง sheet เงินเข้าบริษัท ─────────────────────────────────
    if 'เงินเข้าบริษัท' in wb.sheetnames:
        del wb['เงินเข้าบริษัท']
    ws = wb.create_sheet('เงินเข้าบริษัท', 0)

    all_hdrs = list(erp_headers) + [display for display, _ in WANTED_COLS] + EXTRA_COLS
    for c, h in enumerate(all_hdrs, 1):
        ws.cell(1, c).value = h or ''

    extra_start = N + 1

    YODKHAI_COL  = get_column_letter(extra_start)
    INCOME_COL   = get_column_letter(extra_start + len(WANTED_COLS) - 1)
    cost_per_cn  = extra_start + len(WANTED_COLS)
    cost_tot_cn  = cost_per_cn + 1
    profit_b_cn  = cost_per_cn + 2
    profit_p_cn  = cost_per_cn + 3
    note_cn      = cost_per_cn + 4
    COST_PER_COL = get_column_letter(cost_per_cn)
    COST_TOT_COL = get_column_letter(cost_tot_cn)
    PROFIT_B_COL = get_column_letter(profit_b_cn)

    # ── 5. เขียนแถวข้อมูล ─────────────────────────────────────────────
    log('Writing rows...')
    for i, row_data in enumerate(rows):
        r = i + 2
        for c, val in enumerate(row_data[:N], 1):
            ws.cell(r, c).value = val

        ws.cell(r, 5).value = f'=INDEX(STOCK!$B:$B,MATCH(C{r},STOCK!$I:$I,0))'

        cnt = f'COUNTIF(${F_COL}$2:${F_COL}${CE},{F_COL}{r})'
        sumifs_tpl = "SUMIFS('TikTok'!${c}$2:${c}${e},'TikTok'!${o}$2:${o}${e},{F}{r})"

        for j, (display_name, tiktok_name) in enumerate(WANTED_COLS):
            ec = extra_start + j
            tt_col = tt_col_map.get(tiktok_name)
            if tt_col:
                sumifs = sumifs_tpl.format(c=tt_col, e=TT_END, o=TT_ORDER, F=F_COL, r=r)
                formula = (
                    f'=IF({cnt}>1,'
                    f'{sumifs}/{cnt},'
                    f'{sumifs})'
                )
                ws.cell(r, ec).value = formula
            else:
                ws.cell(r, ec).value = 0
            ws.cell(r, ec).number_format = '#,##0.00'

        # 6 คอลัมน์วิเคราะห์
        ws.cell(r, cost_per_cn).value = (
            f'=IFERROR(IF(${YODKHAI_COL}{r}>0,'
            f'VLOOKUP(I{r},IMPORTRANGE("{COST_URL}","CostORG2!$c$3:$n$90000"),12,0),0),"")'
        )
        ws.cell(r, cost_per_cn).number_format = '#,##0.00'
        ws.cell(r, cost_tot_cn).value = f'={COST_PER_COL}{r}*M{r}'
        ws.cell(r, cost_tot_cn).number_format = '#,##0.00'
        ws.cell(r, profit_b_cn).value = f'={INCOME_COL}{r}+{COST_TOT_COL}{r}'
        ws.cell(r, profit_b_cn).number_format = '#,##0.00'
        ws.cell(r, profit_p_cn).value = (
            f'=IF({YODKHAI_COL}{r}>0,{PROFIT_B_COL}{r}/{YODKHAI_COL}{r},0)'
        )
        ws.cell(r, profit_p_cn).number_format = '0.00%'
        ws.cell(r, note_cn).value = (
            f'=INDEX(IMPORTRANGE("{STOCK_URL}","2026!$I$2:$Z$29000"),'
            f'MATCH(C{r},IMPORTRANGE("{STOCK_URL}","2026!$I$2:$I$29000"),0),18)'
        )

    # ── 6. Format ──────────────────────────────────────────────────────
    log('Formatting...')
    HDR_BLUE  = '0F4267'
    HDR_GREEN = 'E2EFDA'

    for c in range(1, len(all_hdrs) + 1):
        cell = ws.cell(1, c)
        if c <= N:
            cell.fill = PatternFill('solid', fgColor=HDR_BLUE)
            cell.font = Font(color='FFFFFF', bold=True)
        else:
            cell.fill = PatternFill('solid', fgColor=HDR_GREEN)
            cell.font = Font(color='000000', bold=True)
        cell.alignment = Alignment(wrap_text=False, vertical='center')
        ws.column_dimensions[get_column_letter(c)].width = 24

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical='center')

    ws.freeze_panes = 'A2'

    # ── 7. สร้างชีท ใบปะหน้า TIKTOK ──────────────────────────────────
    log('Creating ใบปะหน้า TIKTOK...')

    _m = _re.search(r'(\d{2})(\d{2})\.xlsx$', os.path.basename(OUTPUT))
    title_month = f'{_m.group(2)}/20{_m.group(1)}' if _m else '??/????'

    if 'ใบปะหน้า TIKTOK' in wb.sheetnames:
        del wb['ใบปะหน้า TIKTOK']
    ws_fp = wb.create_sheet('ใบปะหน้า TIKTOK', 1)

    _wl = {name: get_column_letter(extra_start + j)
           for j, (name, _) in enumerate(WANTED_COLS)}

    _REF   = "'เงินเข้าบริษัท'"
    _SH_MA = 'TikTok-MAFIA'
    _SH_TR = 'TikTok-TRC'

    def _sf(col_letter, shop):
        return (f'=SUMIF({_REF}!${SHOP_COL}:${SHOP_COL},"{shop}",'
                f'{_REF}!${col_letter}:${col_letter})')

    _thin = Side(style='thin')
    _bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    FILL_T_TITLE   = PatternFill('solid', fgColor='1F3864')
    FILL_T_HDR     = PatternFill('solid', fgColor='1F3864')
    FILL_T_INC     = PatternFill('solid', fgColor='9DC3E6')
    FILL_T_CAT_RED = PatternFill('solid', fgColor='C00000')
    FILL_T_COST    = PatternFill('solid', fgColor='FFCCCC')
    FILL_T_ORANGE  = PatternFill('solid', fgColor='FAC090')
    FILL_T_CHDR    = PatternFill('solid', fgColor='FAC090')
    _FILL_WHITE    = PatternFill('solid', fgColor='FFFFFF')

    _NUM = '#,##0.00'
    _PCT = '0.00%'

    def _fc(r_n, c_n, val=None, fill=None, bold=False, size=11,
            fmt=None, h='center', color='000000', wrap=False):
        cell = ws_fp.cell(r_n, c_n)
        if val is not None:
            cell.value = val
        if fill:
            cell.fill = fill
        cell.font = Font(bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=h, vertical='center', wrap_text=wrap)
        cell.border = _bdr
        if fmt:
            cell.number_format = fmt
        return cell

    # Row 1: Title
    ws_fp.row_dimensions[1].height = 30
    ws_fp.merge_cells('A1:F1')
    _fc(1, 1, f'สรุปยอดขาย TIKTOK เดือน {title_month}',
        _FILL_WHITE, bold=True, size=14, color='1F3864')

    # Row 2: Headers
    ws_fp.row_dimensions[2].height = 22
    for _ci, _h in enumerate(['หมวด', 'รายการ', 'หมายเหตุ', 'TT-MAFIA', 'TT-TRC', 'รวม'], 1):
        _fc(2, _ci, _h, FILL_T_HDR, bold=True, color='FFFFFF')

    _r = 3

    # ยอดขาย (row 3)
    _fc(_r, 1, 'รายได้', FILL_T_INC, bold=True)
    _fc(_r, 2, 'ยอดขาย', FILL_T_INC, h='left')
    _fc(_r, 3, '', FILL_T_INC)
    _fc(_r, 4, _sf(_wl['ยอดขาย'], _SH_MA), FILL_T_INC, fmt=_NUM, h='right')
    _fc(_r, 5, _sf(_wl['ยอดขาย'], _SH_TR),  FILL_T_INC, fmt=_NUM, h='right')
    _fc(_r, 6, f'=SUM(D{_r}:E{_r})', FILL_T_INC, bold=True, fmt=_NUM, h='right')
    _YODKHAI_R = _r
    _r += 1

    # ค่าใช้จ่าย SHOPEE (10 รายการ)
    _EXPENSE_ITEMS = [
        'ค่าธรรมเนียมคำสั่งซื้อ',
        'ค่าคอมมิชชั่น TikTok Shop',
        'ค่าธรรมเนียมการจัดส่งจริง',
        'ส่วนลดค่าธรรมเนียมการจัดส่งจากแพลตฟอร์ม',
        'ค่าธรรมเนียมการจัดส่งของลูกค้า',
        'เงินสนับสนุนการจัดส่ง',
        'ยอดรวมเงินคืนหลังหักส่วนลดจากผู้ขาย',
        'ค่าธรรมเนียมการจัดส่งสินค้าคืนตามจริง',
        'เงินคืนสำหรับค่าจัดส่ง',
        'ค่าธรรมเนียมสนับสนุนการเติบโตของร้านค้า',
    ]
    _COST_SEC_START = _r
    for _ie, _item in enumerate(_EXPENSE_ITEMS):
        _fc(_r, 1, '' if _ie > 0 else 'placeholder', FILL_T_COST)
        _fc(_r, 2, _item, None, h='left')
        _fc(_r, 3, '', None)
        _col_e = _wl.get(_item, '')
        _fc(_r, 4, _sf(_col_e, _SH_MA) if _col_e else 0, None, fmt=_NUM, h='right')
        _fc(_r, 5, _sf(_col_e, _SH_TR)  if _col_e else 0, None, fmt=_NUM, h='right')
        _fc(_r, 6, f'=SUM(D{_r}:E{_r})', None, fmt=_NUM, h='right')
        _r += 1
    _COST_SEC_END = _r - 1

    ws_fp.merge_cells(f'A{_COST_SEC_START}:A{_COST_SEC_END}')
    _fc(_COST_SEC_START, 1, 'ค่าใช้จ่าย\nSHOPEE', FILL_T_CAT_RED, bold=True, color='FFFFFF', wrap=True)

    # จำนวนเงินเข้าบริษัท
    _fc(_r, 1, '', FILL_T_ORANGE)
    _fc(_r, 2, 'จำนวนเงินเข้าบริษัท', FILL_T_ORANGE, bold=True, h='left')
    _fc(_r, 3, '', FILL_T_ORANGE)
    _fc(_r, 4, f'=SUM(D{_YODKHAI_R}:D{_COST_SEC_END})', FILL_T_ORANGE, bold=True, fmt=_NUM, h='right')
    _fc(_r, 5, f'=SUM(E{_YODKHAI_R}:E{_COST_SEC_END})', FILL_T_ORANGE, bold=True, fmt=_NUM, h='right')
    _fc(_r, 6, f'=SUM(F{_YODKHAI_R}:F{_COST_SEC_END})', FILL_T_ORANGE, bold=True, fmt=_NUM, h='right')
    _INCOME_R = _r
    _r += 1

    # ต้นทุนค่าสินค้า
    _fc(_r, 1, '')
    _fc(_r, 2, 'ต้นทุนค่าสินค้า', h='left')
    _fc(_r, 3, '')
    _fc(_r, 4, _sf(COST_TOT_COL, _SH_MA), fmt=_NUM, h='right')
    _fc(_r, 5, _sf(COST_TOT_COL, _SH_TR),  fmt=_NUM, h='right')
    _fc(_r, 6, f'=SUM(D{_r}:E{_r})', fmt=_NUM, h='right')
    _COSTGOODS_R = _r
    _r += 1

    # ปรับค่าไร้โอตอล
    _fc(_r, 1, '')
    _fc(_r, 2, 'ปรับค่าไร้โอตอล', h='left')
    _fc(_r, 3, '')
    _fc(_r, 4, 0, fmt=_NUM, h='right')
    _fc(_r, 5, 0, fmt=_NUM, h='right')
    _fc(_r, 6, f'=SUM(D{_r}:E{_r})', fmt=_NUM, h='right')
    _ADJ_R = _r
    _r += 1

    # กำไรเบื้องต้น
    _fc(_r, 1, '', FILL_T_INC)
    _fc(_r, 2, 'กำไรเบื้องต้น (บาท)', FILL_T_INC, bold=True, h='left')
    _fc(_r, 3, '', FILL_T_INC)
    _fc(_r, 4, f'=SUM(D{_INCOME_R}:D{_ADJ_R})', FILL_T_INC, bold=True, fmt=_NUM, h='right')
    _fc(_r, 5, f'=SUM(E{_INCOME_R}:E{_ADJ_R})', FILL_T_INC, bold=True, fmt=_NUM, h='right')
    _fc(_r, 6, f'=SUM(F{_INCOME_R}:F{_ADJ_R})', FILL_T_INC, bold=True, fmt=_NUM, h='right')
    _GROSS_R = _r
    _r += 1

    # ต้นทุนค่ากล่อง
    _fc(_r, 1, '')
    _fc(_r, 2, 'ต้นทุนค่ากล่อง', h='left')
    _fc(_r, 3, 0, fmt=_PCT, h='center')
    _fc(_r, 4, f'=D{_GROSS_R}*C{_r}', fmt=_NUM, h='right')
    _fc(_r, 5, f'=E{_GROSS_R}*C{_r}', fmt=_NUM, h='right')
    _fc(_r, 6, f'=SUM(D{_r}:E{_r})', fmt=_NUM, h='right')
    _BOX_R = _r
    _r += 1

    # ต้นทุนค่า PACKING
    _fc(_r, 1, '')
    _fc(_r, 2, 'ต้นทุนค่า PACKING', h='left')
    _fc(_r, 3, 0, fmt=_PCT, h='center')
    _fc(_r, 4, f'=D{_GROSS_R}*C{_r}', fmt=_NUM, h='right')
    _fc(_r, 5, f'=E{_GROSS_R}*C{_r}', fmt=_NUM, h='right')
    _fc(_r, 6, f'=SUM(D{_r}:E{_r})', fmt=_NUM, h='right')
    _PACK_R = _r
    _r += 1

    # กำไรสุทธิ
    _fc(_r, 1, 'กำไรสุทธิ', FILL_T_ORANGE, bold=True)
    _fc(_r, 2, '', FILL_T_ORANGE)
    _fc(_r, 3, '', FILL_T_ORANGE)
    _fc(_r, 4, f'=SUM(D{_GROSS_R}:D{_PACK_R})', FILL_T_ORANGE, bold=True, fmt=_NUM, h='right')
    _fc(_r, 5, f'=SUM(E{_GROSS_R}:E{_PACK_R})', FILL_T_ORANGE, bold=True, fmt=_NUM, h='right')
    _fc(_r, 6, f'=SUM(F{_GROSS_R}:F{_PACK_R})', FILL_T_ORANGE, bold=True, fmt=_NUM, h='right')
    _NET_R = _r
    _r += 1

    # กำไร (%)
    _fc(_r, 1, 'กำไร (%)', FILL_T_ORANGE, bold=True)
    _fc(_r, 2, '', FILL_T_ORANGE)
    _fc(_r, 3, '', FILL_T_ORANGE)
    _fc(_r, 4, f'=D{_NET_R}/D{_YODKHAI_R}', FILL_T_ORANGE, bold=True, fmt=_PCT, h='right')
    _fc(_r, 5, f'=E{_NET_R}/E{_YODKHAI_R}', FILL_T_ORANGE, bold=True, fmt=_PCT, h='right')
    _fc(_r, 6, f'=F{_NET_R}/F{_YODKHAI_R}', FILL_T_ORANGE, bold=True, fmt=_PCT, h='right')
    _r += 2   # เว้น 1 แถว

    # ตารางค่าคอม
    _COM_HDR_MAP = {1: 'ร้านค้า', 3: '%', 4: 'TT-MAFIA', 5: 'TT-TRC',
                    6: 'หักขาดทุน/สินค้าเสียหาย', 7: 'ค่าคอมสุทธิ'}
    for _ci in range(1, 8):
        _fc(_r, _ci, _COM_HDR_MAP.get(_ci, ''), FILL_T_CHDR, bold=True, color='000000')
    _COM_HDR_R = _r
    _r += 1

    # กำไร(ยอดเบิกค่าคอม)
    _fc(_r, 1, 'กำไร(ยอดเบิกค่าคอม)', FILL_T_ORANGE, bold=False, h='left')
    _fc(_r, 2, '', FILL_T_ORANGE)
    _fc(_r, 3, 'ค่าคอม', FILL_T_ORANGE, h='center')
    _fc(_r, 4, f'=D{_NET_R}', FILL_T_ORANGE, fmt=_NUM, h='right')
    _fc(_r, 5, f'=E{_NET_R}', FILL_T_ORANGE, fmt=_NUM, h='right')
    _fc(_r, 6, '', FILL_T_ORANGE)
    _fc(_r, 7, '', FILL_T_ORANGE)
    _PROFIT_COM_R = _r
    _r += 1

    # ค่าคอมแต่ละคน
    _COM_PEOPLE = [('แอร์', 0.10), ('เอ', 0.02), ('พีพี', 0.0)]
    _COM_ROWS = []
    for _name, _rate in _COM_PEOPLE:
        _fc(_r, 1, _name, h='left')
        _fc(_r, 2, '')
        _fc(_r, 3, _rate, fmt=_PCT, h='center')
        _fc(_r, 4, f'=D{_PROFIT_COM_R}*C{_r}', fmt=_NUM, h='right')
        _fc(_r, 5, f'=E{_PROFIT_COM_R}*C{_r}', fmt=_NUM, h='right')
        _fc(_r, 6, 0, fmt=_NUM, h='right')
        _fc(_r, 7, f'=SUM(D{_r}:F{_r})', fmt=_NUM, h='right')
        _COM_ROWS.append(_r)
        _r += 1

    # แถวรวม
    _fc(_r, 1, 'รวม', FILL_T_INC, bold=True, h='left')
    _fc(_r, 2, '', FILL_T_INC)
    _fc(_r, 3, '', FILL_T_INC)
    _fc(_r, 4, f'=SUM(D{_COM_ROWS[0]}:D{_COM_ROWS[-1]})', FILL_T_INC, bold=True, fmt=_NUM, h='right')
    _fc(_r, 5, f'=SUM(E{_COM_ROWS[0]}:E{_COM_ROWS[-1]})', FILL_T_INC, bold=True, fmt=_NUM, h='right')
    _fc(_r, 6, f'=SUM(F{_COM_ROWS[0]}:F{_COM_ROWS[-1]})', FILL_T_INC, bold=True, fmt=_NUM, h='right')
    _fc(_r, 7, f'=SUM(G{_COM_ROWS[0]}:G{_COM_ROWS[-1]})', FILL_T_INC, bold=True, fmt=_NUM, h='right')
    _SUM_ROW = _r
    _r += 1

    # Merge A:B
    for _mr in [_NET_R, _NET_R + 1,
                _COM_HDR_R, _PROFIT_COM_R,
                _COM_ROWS[0], _COM_ROWS[1], _COM_ROWS[2],
                _SUM_ROW]:
        ws_fp.merge_cells(f'A{_mr}:B{_mr}')

    # Column widths
    ws_fp.column_dimensions['A'].width = 24
    ws_fp.column_dimensions['B'].width = 36
    ws_fp.column_dimensions['C'].width = 12
    ws_fp.column_dimensions['D'].width = 18
    ws_fp.column_dimensions['E'].width = 18
    ws_fp.column_dimensions['F'].width = 22
    ws_fp.column_dimensions['G'].width = 18

    log(f'  ใบปะหน้า TIKTOK: {_r - 1} rows')

    # ── 8. บันทึก ──────────────────────────────────────────────────────
    wb.save(OUTPUT)
    log(f'\n✅ {len(rows)} rows → {OUTPUT}')
    return {"rows": len(rows), "output": OUTPUT}


# ─── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python build_tiktok_analysis.py <input_file.xlsx>")
        sys.exit(1)
    main(sys.argv[1])
