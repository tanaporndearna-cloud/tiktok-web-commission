#!/usr/bin/env python3
"""
build_tiktok_erp.py
สร้างไฟล์ Data ERP รายเดือนสำหรับ TikTok Commission ของ TRC Motorsport

Usage (CLI):
  python build_tiktok_erp.py \
    --erp "Data_ERP.xlsx" \
    --abbhtt "ABBHTT05.xlsx" "ABBHTT07.xlsx" \
    --tt_trc "TT-TRC.xlsx" \
    --tt_mafia "TT-MAFIA.xlsx" \
    --stock "stock_full.xlsx" \
    --output "Data ERP 2607.xlsx"

Usage (Python import):
  from scripts.build_tiktok_erp import run_build
  result = run_build(erp_path, abbhtt_paths, tt_trc_path, tt_mafia_path, stock_path, output_path)
"""

import argparse
import sys
import re
import io
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

RED_FILL  = PatternFill("solid", fgColor="FF6666")   # แถวที่เพิ่มจาก ABBHTT
BLUE_FILL = PatternFill("solid", fgColor="99CCFF")   # แถวที่แก้ชื่อร้าน

STOCK_FILE_ID = "1p4-DX-Sq-Mvo_FNAkYtGMkdEBvPDCuPM_sZ_b-_Tm-Y"


# ─── Load helpers ─────────────────────────────────────────────────────────────

def load_stock_from_xlsx(path: str) -> pd.DataFrame:
    """โหลด STOCK จากไฟล์ xlsx (sheet '2026')"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = '2026' if '2026' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.values)
    wb.close()
    data = []
    for r in rows:
        if r is None:
            continue
        row = []
        for v in list(r)[:21]:
            if hasattr(v, 'strftime'):
                row.append(v.strftime('%d/%m/%Y %H:%M:%S'))
            elif v is None:
                row.append('')
            else:
                row.append(str(v))
        data.append(row)
    return pd.DataFrame(data)


def load_erp(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=0, header=None, dtype=str)
    return df.fillna("")


def load_abbhtt(paths: list) -> pd.DataFrame:
    frames = []
    for path in paths:
        df = pd.read_excel(path, sheet_name=0, header=None, dtype=str)
        frames.append(df.fillna(""))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def load_tiktok_sheet(path: str) -> list:
    """โหลดชีท รายละเอียดคำสั่งซื้อ → คืน list of rows"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = next((s for s in wb.sheetnames if 'รายละเอียดคำสั่งซื้อ' in s), wb.sheetnames[0])
    data = list(wb[sn].values)
    wb.close()
    return data


def highlight_row(ws, row_num: int, fill: PatternFill, max_col: int):
    for col in range(1, max_col + 1):
        ws.cell(row=row_num, column=col).fill = fill


def detect_output_filename(erp_df: pd.DataFrame) -> str:
    """ตรวจหาเดือนจากข้อมูล ERP แล้วสร้างชื่อไฟล์ output"""
    import re as _re
    # ลองหาวันที่จากคอลัมน์แรก (col A)
    for i in range(1, min(20, len(erp_df))):
        val = str(erp_df.iloc[i, 0]).strip()
        # รูปแบบ: 01/05/2026 หรือ 2026-05-01
        m = _re.search(r'(\d{2})[/-](\d{2})[/-](\d{4})', val)
        if m:
            dd, mm, yyyy = m.group(1), m.group(2), m.group(3)
            yy = yyyy[2:]
            return f"Data ERP {yy}{mm}.xlsx"
        m2 = _re.search(r'(\d{4})[/-](\d{2})[/-](\d{2})', val)
        if m2:
            yyyy, mm = m2.group(1), m2.group(2)
            yy = yyyy[2:]
            return f"Data ERP {yy}{mm}.xlsx"
    return "Data ERP output.xlsx"


# ─── Main build function ───────────────────────────────────────────────────────

def run_build(erp_path: str, abbhtt_paths: list, tt_trc_path: str,
              tt_mafia_path: str, stock_path: str, output_path: str,
              log_fn=None) -> dict:
    """
    รัน build หลัก — คืน dict สรุปผล
    log_fn: callable(str) สำหรับแสดง log ระหว่างประมวลผล
    """
    def log(msg):
        print(msg)
        if log_fn:
            log_fn(msg)

    log("=" * 60)
    log("TRC Motorsport — TikTok Commission Builder")
    log("=" * 60)

    # ── 1. STOCK ───────────────────────────────────────────────────────────────
    log("\n[1] โหลด STOCK...")
    stock_df = load_stock_from_xlsx(stock_path)
    log(f"  → {stock_path} ({len(stock_df)} แถว)")
    stock_df.columns = range(stock_df.shape[1])

    stock_lookup   = {}  # ABB → order#
    order_to_abb   = {}  # order# → ABB
    for _, row in stock_df.iterrows():
        abb   = str(row[8]).strip() if len(row) > 8 else ""
        order = str(row[1]).strip() if len(row) > 1 else ""
        if abb and abb not in ("", "ABB", " ABB") and order:
            stock_lookup[abb]   = order
            order_to_abb[order] = abb

    log(f"  → STOCK lookup: {len(stock_lookup)} ABB entries")

    # ── 2. Data ERP ────────────────────────────────────────────────────────────
    log("\n[2] โหลด Data ERP...")
    erp_df = load_erp(erp_path)
    erp_rows_raw = erp_df.values.tolist()
    erp_header   = erp_rows_raw[0]
    erp_data     = erp_rows_raw[1:]
    log(f"  → {len(erp_data)} data rows, {erp_df.shape[1]} คอลัมน์")

    # ตรวจหาชื่อไฟล์ output อัตโนมัติ
    auto_filename = detect_output_filename(erp_df)

    # ── 3. ABBHTT ──────────────────────────────────────────────────────────────
    log("\n[3] โหลด ABBHTT...")
    abbhtt_df = load_abbhtt(abbhtt_paths)
    log(f"  → {len(abbhtt_df)} แถว จาก {len(abbhtt_paths)} ไฟล์")

    abbhtt_by_abb = {}
    for _, row in abbhtt_df.iterrows():
        abb = str(row.iloc[2]).strip()
        if abb and abb not in ("", "nan", "เลขที่เอกสาร"):
            abbhtt_by_abb[abb] = list(row)
    log(f"  → ABBHTT lookup (by ABB code): {len(abbhtt_by_abb)} entries")

    # ── 4. TikTok ──────────────────────────────────────────────────────────────
    log("\n[4] โหลด TikTok orders...")
    trc_data   = load_tiktok_sheet(tt_trc_path)
    mafia_data = load_tiktok_sheet(tt_mafia_path)

    tt_headers    = list(trc_data[0]) if trc_data else []
    tt_rows_trc   = [(r, "TikTok-TRC")   for r in trc_data[1:]   if r and any(c for c in r)]
    tt_rows_mafia = [(r, "Tiktok-MAFIA") for r in mafia_data[1:] if r and any(c for c in r)]
    tt_all = tt_rows_trc + tt_rows_mafia
    log(f"  → TRC: {len(tt_rows_trc)}, MAFIA: {len(tt_rows_mafia)}, รวม: {len(tt_all)}")

    order_col_idx = 0  # หมายเลขคำสั่งซื้อ (col แรก)

    # ── 5. สร้าง Workbook ──────────────────────────────────────────────────────
    log("\n[5] สร้างไฟล์ Excel...")
    wb = openpyxl.Workbook()
    erp_max_col = 24  # A–X

    # ── ชีท STOCK ────────────────────────────────────────────────────────────
    ws_stock = wb.active
    ws_stock.title = "STOCK"
    for row in stock_df.values.tolist():
        ws_stock.append([v if str(v) not in ("nan", "") else None for v in row])
    log(f"  → STOCK: {ws_stock.max_row} แถว")

    # ── ชีท Data ERP ─────────────────────────────────────────────────────────
    ws_erp = wb.create_sheet("Data ERP")
    ws_erp.append([v if str(v) not in ("", "nan") else None for v in erp_header])

    added_rows_red  = []
    fixed_rows_blue = []

    for row_data in erp_data:
        row = list(row_data)
        while len(row) < 24:
            row.append(None)

        erp_excel_row = ws_erp.max_row + 1
        ws_erp.append([v if str(v) not in ("", "nan") else None for v in row])

        ws_erp.cell(row=erp_excel_row, column=5).value = (
            f"=INDEX(STOCK!B:B,MATCH('Data ERP'!C{erp_excel_row},STOCK!I:I,0))"
        )
        product_code = str(row[8]).strip() if len(row) > 8 else ""
        if product_code.startswith("BP051"):
            ws_erp.cell(row=erp_excel_row, column=6).value = 0
        else:
            abb_code  = str(row[2]).strip()
            order_val = stock_lookup.get(abb_code, "")
            ws_erp.cell(row=erp_excel_row, column=6).value = order_val if order_val else None

    log(f"  → Data ERP: {ws_erp.max_row - 1} data rows")

    erp_order_to_excel_row = {}
    for er in range(2, ws_erp.max_row + 1):
        val = ws_erp.cell(row=er, column=6).value
        if val:
            erp_order_to_excel_row[str(val).strip()] = er

    # ── ชีท ABBHTT ───────────────────────────────────────────────────────────
    ws_abbhtt = wb.create_sheet("ABBHTT")
    for row in abbhtt_df.values.tolist():
        ws_abbhtt.append([v if str(v) not in ("", "nan") else None for v in row])
    log(f"  → ABBHTT: {ws_abbhtt.max_row} แถว")

    # ── ชีท TikTok ───────────────────────────────────────────────────────────
    ws_tt = wb.create_sheet("TikTok")
    full_headers = [None] + list(tt_headers) + ["SHOP QC"]
    ws_tt.append(full_headers)
    shopqc_col = len(full_headers)

    pending_abbhtt = []

    for tt_excel_row, (row, shop_name) in enumerate(tt_all, 2):
        row_vals  = list(row)
        order_raw = str(row_vals[order_col_idx]).strip() if row_vals[order_col_idx] else ""
        order_no  = re.sub(r"[^\d]", "", order_raw)

        full_row = [shop_name] + row_vals
        ws_tt.append([v if v not in (None, "") or isinstance(v, (int, float)) else None for v in full_row])

        if not order_no:
            continue

        in_erp = order_no in erp_order_to_excel_row

        if not in_erp:
            abb = order_to_abb.get(order_no)
            if abb and abb in abbhtt_by_abb:
                pending_abbhtt.append((order_no, shop_name, list(abbhtt_by_abb[abb])))
                erp_order_to_excel_row[order_no] = -1
            # ถ้าไม่พบใน ABBHTT → ข้ามไป (ไม่เพิ่ม ยกเลิกก่อนส่ง ตาม SKILL)

    # เพิ่มแถว ABBHTT (แดง)
    for order_no, shop_name, abbhtt_row in pending_abbhtt:
        while len(abbhtt_row) < 24:
            abbhtt_row.append(None)

        new_erp_row = ws_erp.max_row + 1
        ws_erp.append([v if str(v) not in ("", "nan") else None for v in abbhtt_row])

        ws_erp.cell(row=new_erp_row, column=5).value = (
            f"=INDEX(STOCK!B:B,MATCH('Data ERP'!C{new_erp_row},STOCK!I:I,0))"
        )
        new_product_code = str(abbhtt_row[8]).strip() if len(abbhtt_row) > 8 else ""
        if new_product_code.startswith("BP051"):
            ws_erp.cell(row=new_erp_row, column=6).value = 0
        else:
            ws_erp.cell(row=new_erp_row, column=6).value = order_no

        current_shop = str(abbhtt_row[22]).strip() if len(abbhtt_row) > 22 and abbhtt_row[22] else ""
        if current_shop != shop_name:
            ws_erp.cell(row=new_erp_row, column=23).value = shop_name
            fixed_rows_blue.append(new_erp_row)

        highlight_row(ws_erp, new_erp_row, RED_FILL, erp_max_col)
        added_rows_red.append(new_erp_row)
        erp_order_to_excel_row[order_no] = new_erp_row

    # ใส่สูตร VLOOKUP ใน SHOP QC
    for tt_excel_row in range(2, ws_tt.max_row + 1):
        ws_tt.cell(row=tt_excel_row, column=shopqc_col).value = (
            f"=VLOOKUP(B{tt_excel_row},'Data ERP'!E:W,19,0)"
        )

    # ตรวจ SHOP QC ≠ col A → ไฮไลท์ฟ้าใน ERP
    for tt_excel_row, (row, shop_name) in enumerate(tt_all, 2):
        order_raw = str(row[order_col_idx]).strip() if row[order_col_idx] else ""
        order_no  = re.sub(r"[^\d]", "", order_raw)
        er = erp_order_to_excel_row.get(order_no)
        if er and er not in added_rows_red:
            erp_shop = ws_erp.cell(row=er, column=23).value
            erp_shop = str(erp_shop).strip() if erp_shop else ""
            if erp_shop and erp_shop != shop_name:
                ws_erp.cell(row=er, column=23).value = shop_name
                if er not in fixed_rows_blue:
                    fixed_rows_blue.append(er)
                    highlight_row(ws_erp, er, BLUE_FILL, erp_max_col)

    blue_only = [r for r in set(fixed_rows_blue) if r not in added_rows_red]
    log(f"\n[6] Highlights:")
    log(f"  → แดง (#FF6666) เพิ่มจาก ABBHTT: {len(added_rows_red)} แถว")
    log(f"  → ฟ้า (#99CCFF) แก้ชื่อร้าน:     {len(blue_only)} แถว")

    wb.save(output_path)
    log(f"\n✅ บันทึก: {output_path}")

    return {
        "output_filename": auto_filename,
        "tt_rows": len(tt_all),
        "added_red": len(added_rows_red),
        "fixed_blue": len(blue_only),
    }


# ─── CLI ──────────────────────────────────────────────────────────────────────
def build(args):
    result = run_build(
        erp_path=args.erp,
        abbhtt_paths=args.abbhtt,
        tt_trc_path=args.tt_trc,
        tt_mafia_path=args.tt_mafia,
        stock_path=args.stock,
        output_path=args.output,
    )
    print(f"TikTok rows: {result['tt_rows']}")
    print(f"Added ERP (red):  {result['added_red']}")
    print(f"Fixed shop (blue): {result['fixed_blue']}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="TRC TikTok Commission Builder")
    parser.add_argument("--erp",      required=True,            help="Data ERP Excel file")
    parser.add_argument("--abbhtt",   required=True, nargs="+", help="ABBHTT Excel file(s)")
    parser.add_argument("--tt_trc",   required=True,            help="TT-TRC Excel file")
    parser.add_argument("--tt_mafia", required=True,            help="TT-MAFIA Excel file")
    parser.add_argument("--stock",    required=True,            help="STOCK xlsx")
    parser.add_argument("--output",   default="Data ERP output.xlsx")
    args = parser.parse_args()
    build(args)
